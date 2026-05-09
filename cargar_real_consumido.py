#!/usr/bin/env python3
"""Cargador de Real Consumido — rellena archivos de planificación con montos reales de facturas."""

import io, sys, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from pathlib import Path

BASE   = Path(r"c:\Users\franco.faustin\OneDrive - OneWorkplace\Escritorio\Claude Van Damme\Diferenciador de facturas")
RC_DIR = BASE / "Archivos Real Consumido"
NUM_FMT = "#,##0.00"
YT_RATE = 0.032
DV_RATE = 0.085

def tf(amount, rate): return round(amount * rate / (1 + rate), 2)

# ── Datos de facturas por anunciante ─────────────────────────────────────────

INVOICES = {
    "Fargo": {
        "meta":    {"subtotal": 9_209_265.69},
        "formats": [
            {"label": "DV360 Display",  "amount": 9_959_121.85,  "rate": DV_RATE, "kws": ["display"]},
            {"label": "YouTube Bumper", "amount": 15_623_389.46, "rate": YT_RATE, "kws": ["bumper"]},
        ],
        "cm360":   {"label": "Adserver (CM360)", "subtotal": 710_422.72},
    },
    "Artesano": {
        "meta":    {"subtotal": 47_399_513.87},
        "tiktok":  {"subtotal": 12_375_285.58},
        "formats": [
            {"label": "YouTube Trueview", "amount": 20_657_431.19, "rate": YT_RATE, "kws": ["trueview","true view"]},
            {"label": "YouTube Bumper",   "amount": 20_637_703.36, "rate": YT_RATE, "kws": ["bumper"]},
        ],
        "cm360":   {"label": "Adserver (CM360)", "subtotal": 374_278.02},
    },
    "Oroweat": {
        "meta":    {"subtotal": 57_016_985.85},
        "formats": [
            {"label": "YouTube Trueview", "amount": 18_234_214.23, "rate": YT_RATE, "kws": ["trueview","true view"]},
            {"label": "YouTube Bumper",   "amount": 14_175_229.13, "rate": YT_RATE, "kws": ["bumper"]},
        ],
        "cm360":   {"label": "Adserver (CM360)", "subtotal": 567_280.31},
    },
    "Rapiditas": {
        "meta":    {"subtotal": 26_946_898.81},
        "tiktok":  {"subtotal": 14_654_677.66},
        "formats": [
            {"label": "YouTube Trueview", "amount": 18_208_813.31, "rate": YT_RATE, "kws": ["trueview","true view","video"]},
            {"label": "DV360 Display",    "amount": 10_090_364.22, "rate": DV_RATE, "kws": ["display"]},
        ],
        "cm360":   {"label": "Adserver (CM360)", "subtotal": 339_387.11},
    },
    "Salmas": {
        "meta":    {"subtotal": 45_547_658.73},
        "formats": [
            {"label": "YouTube Trueview", "amount": 12_237_607.80, "rate": YT_RATE, "kws": ["trueview","true view","video"]},
            {"label": "YouTube Bumper",   "amount": 12_214_397.06, "rate": YT_RATE, "kws": ["bumper"]},
        ],
        "cm360":   {"label": "Adserver (CM360)", "subtotal": 209_041.76},
    },
    "Bimbo_BTS": {
        # Meta: solo campañas Back to School de la factura 252247522
        "meta":  {"subtotal": 8_747_557.67},
        "formats": [
            {"label": "DV360 Display", "amount": 3_234_377.12, "rate": DV_RATE, "kws": ["display","programm"]},
        ],
        "cm360": {"label": "Adserver (CM360)", "subtotal": 102_569.21},
    },
    "Bimbo_Barcelona": {
        # Meta: solo campañas Barcelona de la factura 252247522
        "meta": {"subtotal": 1_571_505.14},
    },
}

FILE_MAP = {
    "Real Consumido_Fargo_Q1_2026_Omnet_V5.xlsx":         "Fargo",
    "Real Consumido_Artesano_Reloaded Mar-Jun.xlsx":       "Artesano",
    "Real Consumido_Oroweat_V4 - Omnet.xlsx":             "Oroweat",
    "Real Consumido_Rapiditas_CampañaRegional_Feb-Mar.xlsx": "Rapiditas",
    "Real Consumido_Salmas_V7.xlsx":                      "Salmas",
    "Real Consumido_BackToSchol_Feb-Abr.xlsx":            "Bimbo_BTS",
    "Real Consumido_Barcelona_CampañaGlobal_EneJul.xlsx": "Bimbo_Barcelona",
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def get_sheet(wb):
    for name in wb.sheetnames:
        if "digital" in name.lower():
            return wb[name]
    return wb.active

def find_april_block(ws):
    """Return (inversion_row, inversion_col) for Abril."""
    abril_col = None

    # Priority 1: merged cell in row 13
    for m in ws.merged_cells.ranges:
        if m.min_row == 13:
            val = str(ws.cell(m.min_row, m.min_col).value or "").upper()
            if "ABRIL" in val or ("ABR" in val and "ABRI" not in val):
                abril_col = m.min_col
                break

    # Priority 2: any cell rows 13-16 exactly 'ABRIL' or 'ABR'
    if not abril_col:
        for row in range(13, 17):
            for col in range(1, ws.max_column + 1):
                val = str(ws.cell(row, col).value or "").strip().upper()
                if val in ("ABRIL", "15 DE ABRIL", "15 ABRIL", "ABR"):
                    abril_col = col
                    break
            if abril_col:
                break

    if not abril_col:
        return None, None

    # Find INVERSION starting at abril_col in rows 15-16
    for row in [15, 16]:
        for col in range(abril_col, min(abril_col + 4, ws.max_column + 1)):
            val = str(ws.cell(row, col).value or "").upper()
            if "INVERSION" in val or "INVERSIÓN" in val:
                return row, col

    # Fallback: INVERSION is at abril_col itself
    return 15, abril_col

def unmerge_if_needed(ws, row, col):
    """If cell is part of a merged range, unmerge it."""
    target = (row, col)
    for m in list(ws.merged_cells.ranges):
        if m.min_row <= row <= m.max_row and m.min_col <= col <= m.max_col:
            ws.unmerge_cells(str(m))
            return True
    return False

def ensure_rc_col(ws, inv_row, inv_col):
    """Return REAL CONSUMIDO column (inserting if necessary)."""
    rc_col = inv_col + 1

    # Unmerge if the target cell is part of a merged range
    unmerge_if_needed(ws, inv_row, rc_col)

    next_val = str(ws.cell(inv_row, rc_col).value or "").upper()
    if "REAL" in next_val:
        print(f"    REAL CONSUMIDO ya existe en {get_column_letter(rc_col)}{inv_row}")
        return rc_col
    if ws.cell(inv_row, rc_col).value is None:
        # Empty — just write header
        pass
    else:
        # Occupied — insert column
        print(f"    Insertando columna en {get_column_letter(rc_col)} (antes: '{ws.cell(inv_row, rc_col).value}')")
        ws.insert_cols(rc_col)
    ws.cell(inv_row, rc_col).value = "REAL CONSUMIDO"
    ws.cell(inv_row, rc_col).font = Font(bold=True)
    return rc_col

def find_row(ws, keywords, start=16, col_range=(2, 9)):
    """Find first row where any cell matches keywords (case-insensitive)."""
    for row in range(start, min(ws.max_row + 1, start + 50)):
        for col in range(col_range[0], col_range[1] + 1):
            val = str(ws.cell(row, col).value or "").lower()
            if any(k in val for k in keywords):
                return row
    return None

def write_rc(ws, row, col, value):
    ws.cell(row, col).value = value
    ws.cell(row, col).number_format = NUM_FMT

def insert_label_row(ws, after_row, label, label_col, rc_col, amount):
    """Insert a new row below after_row with a label and RC amount."""
    ws.insert_rows(after_row + 1)
    ws.cell(after_row + 1, label_col).value = label
    ws.cell(after_row + 1, label_col).font = Font(italic=True)
    ws.cell(after_row + 1, rc_col).value = amount
    ws.cell(after_row + 1, rc_col).number_format = NUM_FMT
    ws.cell(after_row + 1, rc_col).font = Font(italic=True)
    return after_row + 1  # return new row number

# ── Main processor ────────────────────────────────────────────────────────────

def process(xlsx_path: Path, adv_key: str):
    inv_data = INVOICES.get(adv_key)
    if not inv_data:
        print(f"  ⚠ Sin datos para {adv_key}")
        return

    # Backup
    bak = xlsx_path.with_name(xlsx_path.stem + "_BACKUP.xlsx")
    if not bak.exists():
        shutil.copy2(xlsx_path, bak)
        print(f"  Backup: {bak.name}")

    wb = openpyxl.load_workbook(xlsx_path, keep_links=False)
    ws = get_sheet(wb)
    print(f"  Hoja: '{ws.title}'")

    inv_row, inv_col = find_april_block(ws)
    if not inv_col:
        print(f"  ✗ No se encontró bloque de Abril")
        return
    print(f"  INVERSION Abril: {get_column_letter(inv_col)}{inv_row}")

    rc_col = ensure_rc_col(ws, inv_row, inv_col)
    print(f"  REAL CONSUMIDO col: {get_column_letter(rc_col)}")

    data_start = inv_row + 1
    label_col = 4   # col D for inserted labels
    # After finding Meta/TikTok rows, DV360 format search starts from here
    format_search_start = data_start

    # Collect all insertions (done bottom-to-top at the end)
    insertions = []  # (after_row, label, amount)

    # ── Meta ─────────────────────────────────────────────────────────────────
    if "meta" in inv_data:
        meta_row = find_row(ws, ["facebook"], data_start)
        if meta_row:
            write_rc(ws, meta_row, rc_col, inv_data["meta"]["subtotal"])
            print(f"  ✓ Meta fila {meta_row}: {inv_data['meta']['subtotal']:,.2f}")
            iibb_amt = round(inv_data["meta"]["subtotal"] * 0.02, 2)
            insertions.append((meta_row, "IIBB Meta (2%)", iibb_amt))
            format_search_start = max(format_search_start, meta_row + 1)
        else:
            print(f"  ⚠ No se encontró fila Meta (Facebook)")

    # ── TikTok ───────────────────────────────────────────────────────────────
    if "tiktok" in inv_data:
        tt_row = find_row(ws, ["tiktok"], data_start)
        if tt_row:
            write_rc(ws, tt_row, rc_col, inv_data["tiktok"]["subtotal"])
            print(f"  ✓ TikTok fila {tt_row}: {inv_data['tiktok']['subtotal']:,.2f}")
            format_search_start = max(format_search_start, tt_row + 1)
        else:
            print(f"  ⚠ No se encontró fila TikTok (puede no estar planificado)")

    # ── DV360 formats ────────────────────────────────────────────────────────
    last_fmt_row = data_start
    if "formats" in inv_data:
        for fmt in inv_data["formats"]:
            fmt_row = find_row(ws, fmt["kws"], format_search_start)
            if fmt_row:
                write_rc(ws, fmt_row, rc_col, fmt["amount"])
                tf_amt = tf(fmt["amount"], fmt["rate"])
                print(f"  ✓ {fmt['label']} fila {fmt_row}: {fmt['amount']:,.2f}  (TF: {tf_amt:,.2f})")
                insertions.append((fmt_row, f"Tech Fee — {fmt['label']}", tf_amt))
                if fmt_row > last_fmt_row:
                    last_fmt_row = fmt_row
            else:
                print(f"  ⚠ No se encontró fila para {fmt['label']} {fmt['kws']}")

    # ── CM360 ─────────────────────────────────────────────────────────────────
    if "cm360" in inv_data:
        cm_row = find_row(ws, ["adserver", "cm360", "ad serving", "adserving"], data_start)
        if cm_row:
            write_rc(ws, cm_row, rc_col, inv_data["cm360"]["subtotal"])
            print(f"  ✓ CM360 fila {cm_row}: {inv_data['cm360']['subtotal']:,.2f}")
        else:
            # Insert after last format row
            insertions.append((last_fmt_row, inv_data["cm360"]["label"], inv_data["cm360"]["subtotal"]))
            print(f"  + CM360: insertando después de fila {last_fmt_row}: {inv_data['cm360']['subtotal']:,.2f}")

    # ── Insert rows (bottom → top to keep row numbers stable) ─────────────────
    # Same row: non-TF items first so TF ends up right after its platform row
    insertions.sort(key=lambda x: (-x[0], "Tech Fee" in x[1]))
    for after_row, label, amount in insertions:
        new_row = insert_label_row(ws, after_row, label, label_col, rc_col, amount)
        print(f"  + Fila {new_row}: '{label}' → {amount:,.2f}")

    wb.save(xlsx_path)
    print(f"  ✓ Guardado: {xlsx_path.name}")

# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    print("=" * 65)
    print("  CARGADOR DE REAL CONSUMIDO — Abril 2026")
    print("=" * 65)

    cells_updated = 0
    files_modified = []

    for fname, adv_key in FILE_MAP.items():
        path = RC_DIR / fname
        if not path.exists():
            print(f"\n⚠ Archivo no encontrado: {fname}")
            continue
        print(f"\n{'─'*65}")
        print(f"ARCHIVO : {fname}")
        print(f"Anunciante: {adv_key}")
        try:
            process(path, adv_key)
            files_modified.append(fname)
        except Exception as e:
            print(f"  ✗ ERROR: {e}")
            import traceback; traceback.print_exc()

    # Takis
    print(f"\n{'─'*65}")
    print("TAKIS: campaña hasta Marzo. Sin columna de Abril.")
    print("  → Monto a registrar (CM360): ARS 38.34 — archivo NO modificado.")

    print(f"\n{'='*65}")
    print(f"✅ Archivos procesados: {len(files_modified)}")
    print(f"📋 Archivos modificados:")
    for f in files_modified:
        print(f"   • {f}")
    print(f"⚠  Sin modificar: Takis (sin columna Abril)")
    print("=" * 65)

if __name__ == "__main__":
    main()
