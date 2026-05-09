#!/usr/bin/env python3
"""Procesador de facturas publicitarias — Google DV360/CM360, Meta, TikTok."""

import csv as csv_module
import io
import re
import shutil
import sys
from pathlib import Path

# Fix Windows console encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
import pdfplumber
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE       = Path(r"c:\Users\franco.faustin\OneDrive - OneWorkplace\Escritorio\Claude Van Damme\Diferenciador de facturas")
ENTRADA    = BASE / "facturas_entrada"
PROCESADAS = BASE / "facturas_procesadas"
REPORTES   = BASE / "reportes"
SPENT_PATH = BASE / "Spent Google" / "Spent por campaign.xlsx"

# ── Helpers ────────────────────────────────────────────────────────────────────

def pdf_text(path: Path) -> str:
    pages = []
    with pdfplumber.open(path) as pdf:
        for pg in pdf.pages:
            t = pg.extract_text()
            if t:
                pages.append(t)
    return "\n".join(pages)


def parse_num(s: str) -> float:
    """Parse ARS amounts: English 32,409,443.35 or Argentine 32.409.443,35."""
    if not s:
        return 0.0
    s = str(s).strip()
    neg = s.startswith("-")
    s = s.lstrip("-").replace("ARS", "").replace("$", "").strip()
    last_dot   = s.rfind(".")
    last_comma = s.rfind(",")
    if last_comma > last_dot:
        s = s.replace(".", "").replace(",", ".")   # Argentine
    else:
        s = s.replace(",", "")                     # English
    try:
        v = float(s)
        return -v if neg else v
    except ValueError:
        return 0.0


def advertiser_from(s: str):
    u = s.upper()
    if "OROWEAT" in u or "_ORO_" in u or "ORO_ARG" in u:
        return "Oroweat"
    if "ARTESANO" in u or "_ART_" in u:
        return "Artesano"
    if "RAPIDITAS" in u or "_RPD_" in u:
        return "Rapiditas"
    if "SALMAS" in u or "_SAL_" in u or "SLMS" in u:
        return "Salmas"
    if "_FAR_" in u or "FARGO" in u or "FRGP" in u:
        return "Fargo"
    if "TIA ROSA" in u or "TIAROSA" in u or "TIA_ROSA" in u:
        return "Tia Rosa"
    if "TAKIS" in u or "_TAK_" in u:
        return "Takis"
    if "ZENISSIMO" in u:  # BimboZenissimo_ARG = Salmas
        return "Salmas"
    if "BIMBO" in u or "BACK TO SCHOOL" in u or "BBAR" in u or "_BIM_BIM_" in u:
        return "Bimbo"
    return None


def meta_format(campaign: str) -> str:
    c = campaign.upper()
    if re.search(r"-TR(?:[||\s]|$)|_TR_", c):
        return "Meta Tráfico"
    if re.search(r"_AWA_|-AW(?:[||\s]|$)", c):
        return "Meta Awareness"
    return "Meta"


def cm360_format(desc: str, unit: str) -> str:
    d = desc.upper()
    u = (unit or "").upper()
    if "SEGUIMIENTO DE CLICS" in d or u == "CPC":
        return "Adserving - CPC"
    if "PUBLICACIÓN ANUNCIOS GRÁFICOS" in d or "PUBLICACION ANUNCIOS GRAFICOS" in d:
        return "Display Programático"
    if "SEGUIMIENTO DE IMPRESIONES" in d or u == "CPM":
        if any(x in d for x in ["_DD_", "_MMD_", "DSPL", "_DIS_", "MEDIOSDIR", "MEDIOS DIR", "DISPLAY"]):
            return "Adserving Display - CPM"
        if any(x in d for x in ["_VID_", "_YT_", "_OLV_", "VIDEO", "YOUTUBE", "YT-"]):
            return "Adserving Video - CPM"
        return "Adserving Display - CPM"
    return "Adserving Display - CPM"


def format_tech_fee(formato: str, amount: float) -> float:
    """Tech Fee = subtotal × rate / (1 + rate). YouTube=3.2%, DV360=8.5%."""
    rate = 0.032 if "YouTube" in formato else 0.085
    return round(amount * rate / (1 + rate), 2)


def io_format(io_name: str) -> str:
    """Detect ad format from DV360 Insertion Order name."""
    io = io_name.upper()
    if any(x in io for x in ["MASTHEAD", "MSTH"]):
        return "YouTube Masthead"
    if any(x in io for x in ["TRV", "TRUEVIEW", "STTV"]):
        return "YouTube Trueview"
    if any(x in io for x in ["BUM", "BUMPER", "BUMP"]):
        return "YouTube Bumper"
    if ("DISPLAY" in io or "DISP" in io) and ("DV3" in io or "D3DP" in io):
        return "DV360 Display"
    if "YOU" in io or "YTID" in io or "YTIS" in io or "YTMP" in io:
        return "YouTube Trueview"
    if "VID" in io or "VIDE" in io:
        return "DV360 Video"
    return "DV360 Display"


def load_spent_data() -> dict:
    """Load Spent por campaign.xlsx → {advertiser: [{campaign, formato, amount, advertiser}]}"""
    if not SPENT_PATH.exists():
        print(f"  ⚠ Archivo Spent no encontrado: {SPENT_PATH}")
        return {}
    wb = openpyxl.load_workbook(SPENT_PATH, data_only=True)
    ws = wb.active
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 4:
            continue
        campaign, io_name, _, amount = row[0], row[1], row[2], row[3]
        if not campaign or not io_name or not isinstance(amount, (int, float)):
            continue
        adv = advertiser_from(str(campaign))
        if not adv:
            continue
        fmt = io_format(str(io_name))
        if adv not in result:
            result[adv] = []
        result[adv].append({
            "campaign": str(campaign),
            "formato": fmt,
            "amount": round(float(amount), 2),
            "advertiser": adv,
        })
    return result


def tiktok_format(campaign: str) -> str:
    c = campaign.upper()
    if "TOPVIEW" in c:
        return "TopView"
    if "IN-FEED" in c or "IN_FEED" in c:
        return "In-Feed"
    if "BRANDED" in c:
        return "Branded Effect"
    return "Video"


# ── Google CSV helpers ─────────────────────────────────────────────────────────

def find_csv(invoice_num: str):
    for d in (ENTRADA / "Google").iterdir():
        if d.is_dir():
            c = d / f"{invoice_num}.csv"
            if c.exists():
                return c
    return None


def parse_google_csv(csv_path: Path, is_dv360: bool):
    lines = []
    advertisers = set()
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv_module.reader(f))

    # Locate header row
    hdr_idx = None
    for i, row in enumerate(rows):
        if "Descripción" in ",".join(row) or "Descripcion" in ",".join(row):
            hdr_idx = i
            break
    if hdr_idx is None:
        return lines, advertisers

    hdr = rows[hdr_idx]
    col_desc = col_unit = col_amount = None
    for j, h in enumerate(hdr):
        h = h.strip()
        if "Descripción" in h or "Descripcion" in h:
            col_desc = j
        elif "Unidad" in h:
            col_unit = j
        elif "Importe" in h:
            col_amount = j

    if col_desc is None or col_amount is None:
        return lines, advertisers

    for row in rows[hdr_idx + 1:]:
        if not row:
            continue
        desc     = row[col_desc].strip()   if col_desc   < len(row) else ""
        unit     = row[col_unit].strip()   if col_unit   is not None and col_unit < len(row) else ""
        amount_s = row[col_amount].strip() if col_amount < len(row) else ""

        if not desc or not amount_s:
            continue
        if any(x in desc.upper() for x in ["IVA", "I.V.A.", "PERCEP", "INGRESOS BRUTOS", "RÉGIMEN", "REGIMEN"]):
            continue

        amount = parse_num(amount_s)

        if is_dv360:
            m = re.search(r"Anunciante:\s+([\w-]+)", desc)
            adv_raw = m.group(1) if m else desc
            adv = advertiser_from(adv_raw) or advertiser_from(desc) or "Desconocido"
            fmt = "Costo de medios" if "Costo de medios" in desc else "Tech Fee"
            campaign = adv_raw
        else:
            m = re.search(r'Campaña:\s+"([^"]+)"', desc)
            campaign = m.group(1) if m else desc
            m2 = re.search(r'Anunciante:\s+"([^"]+)"', desc)
            adv_raw = m2.group(1) if m2 else campaign
            adv = advertiser_from(adv_raw) or advertiser_from(campaign) or "Desconocido"
            fmt = cm360_format(desc + " " + campaign, unit)

        if adv != "Desconocido":
            advertisers.add(adv)
        lines.append({"campaign": campaign, "formato": fmt, "amount": amount, "advertiser": adv})

    return lines, advertisers


# ── Invoice parsers ────────────────────────────────────────────────────────────

def parse_google(pdf_path: Path) -> dict:
    text = pdf_text(pdf_path)
    inv = {"plataforma": "Google", "source_path": pdf_path}

    m = re.search(r"Número de documento[:\s.]+(\d+)", text)
    inv["numero"] = m.group(1) if m else pdf_path.stem

    is_dv360 = bool(re.search(r"Display.{0,10}Video\s*360|DV[-\s]?360", text, re.IGNORECASE))
    inv["subcategoria"] = "DV360" if is_dv360 else "CM360"

    m = re.search(r"Subtotal en ARS\s+ARS\s+([\d,]+\.\d{2})", text)
    inv["subtotal"] = parse_num(m.group(1)) if m else 0.0

    csv_path = find_csv(inv["numero"])
    lines, advertisers = (parse_google_csv(csv_path, is_dv360) if csv_path else ([], set()))

    # PDF fallback for DV360 when CSV unavailable
    if not lines and is_dv360:
        for m in re.finditer(
            r"(Costo de medios|Tarifa de plataforma).*?Anunciante:\s*(\w+).*?1\s+EA\s+([\d,]+\.\d{2})",
            text, re.DOTALL,
        ):
            tipo, adv_raw, amount_s = m.group(1), m.group(2), m.group(3)
            adv = advertiser_from(adv_raw) or adv_raw
            advertisers.add(adv)
            fmt = "Costo de medios" if "Costo de medios" in tipo else "Tech Fee"
            lines.append({"campaign": adv_raw, "formato": fmt, "amount": parse_num(amount_s), "advertiser": adv})

    if len(advertisers) > 1:
        inv["anunciante"] = "MultiAnunciante"
    elif advertisers:
        inv["anunciante"] = list(advertisers)[0]
    else:
        m2 = re.search(r'Anunciante[:\s]+"?(\w+)_ARG', text)
        if m2:
            inv["anunciante"] = advertiser_from(m2.group(1)) or m2.group(1)
        else:
            inv["anunciante"] = "Desconocido"

    inv["lineas"] = lines
    return inv


def parse_meta(pdf_path: Path) -> dict:
    text = pdf_text(pdf_path)
    inv = {"plataforma": "Meta", "subcategoria": "", "source_path": pdf_path}

    m = re.search(r"Invoice #:\s*(\d+)", text)
    inv["numero"] = m.group(1) if m else pdf_path.stem

    m = re.search(r"Subtotal:\s*([\d,]+\.\d{2})", text)
    inv["subtotal"] = parse_num(m.group(1)) if m else 0.0

    lines = []
    advertisers = set()

    # Use pdfplumber table extraction (handles wrapped cell content)
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for tbl in (page.extract_tables() or []):
                for row in tbl:
                    if not row:
                        continue
                    line_num = str(row[0] or "").strip()
                    if not line_num.isdigit():
                        continue
                    # Campaign is in col 1, amount in last col
                    campaign_raw = str(row[1] or "").strip()
                    amount_s     = str(row[-1] or "").strip()
                    # Join hyphen-continuations first, then normalize spaces
                    campaign = re.sub(r"-\n", "-", campaign_raw)
                    campaign = " ".join(campaign.split())
                    if not campaign or any(
                        x in campaign.lower() for x in ["subtotal", "invoice total", "vat", "freight"]
                    ):
                        continue
                    amount = parse_num(amount_s)
                    adv = advertiser_from(campaign) or "Desconocido"
                    if adv != "Desconocido":
                        advertisers.add(adv)
                    fmt = meta_format(campaign)
                    lines.append({"campaign": campaign, "formato": fmt, "amount": amount, "advertiser": adv})

    # Fallback: text regex if table extraction gave nothing
    if not lines:
        for m in re.finditer(r"^\s*(\d+)\s+(.*?)\s+([-\d,]+\.\d{2})\s*$", text, re.MULTILINE):
            campaign = " ".join(m.group(2).split())
            if any(x in campaign.lower() for x in ["subtotal", "total", "vat", "freight", "description"]):
                continue
            amount = parse_num(m.group(3))
            adv = advertiser_from(campaign) or "Desconocido"
            if adv != "Desconocido":
                advertisers.add(adv)
            lines.append({"campaign": campaign, "formato": meta_format(campaign), "amount": amount, "advertiser": adv})

    inv["anunciante"] = (
        "MultiAnunciante" if len(advertisers) > 1
        else (list(advertisers)[0] if advertisers else "Desconocido")
    )
    # Fallback: assign invoice-level advertiser to lines with no brand
    inv_adv = inv["anunciante"]
    for line in lines:
        if line["advertiser"] == "Desconocido" and inv_adv not in ("Desconocido", "MultiAnunciante"):
            line["advertiser"] = inv_adv

    # IIBB: 2% sobre el subtotal de la factura Meta
    inv["iibb"] = round(inv["subtotal"] * 0.02, 2)
    lines.append({
        "campaign": "IIBB 2% — Meta",
        "formato": "IIBB Meta",
        "amount": inv["iibb"],
        "advertiser": inv_adv,
    })

    inv["lineas"] = lines
    return inv


def parse_tiktok(pdf_path: Path) -> dict:
    text = pdf_text(pdf_path)
    inv = {"plataforma": "TikTok", "subcategoria": "", "source_path": pdf_path}

    m = re.search(r"Factura\s+(FCA[-\d]+)", text)
    if not m:
        m = re.search(r"(FCA-\d{4}-\d+)", text)
    inv["numero"] = m.group(1) if m else pdf_path.stem

    # Subtotal in Argentine format (dots=thousands, comma=decimal)
    m = re.search(r"Subtotal\s+ARS\s+([\d.]+,\d{2})", text)
    inv["subtotal"] = parse_num(m.group(1)) if m else 0.0

    # Advertiser: from campaign name (more reliable than "Anunciante" field)
    m = re.search(r"(Rapiditas|Oroweat|Artesano|Bimbo|Salmas|Fargo)_\d{4}", text, re.IGNORECASE)
    inv["anunciante"] = m.group(1).capitalize() if m else "Desconocido"

    # Campaign name
    m = re.search(r"(\w+_2026_BIM_[A-Z_]+ARG[^\s\n]+)", text)
    if not m:
        m = re.search(r"(\w+_2026_[^\n]+)", text)
    campaign = " ".join(m.group(1).split()) if m else inv["anunciante"]
    fmt = tiktok_format(campaign)

    inv["lineas"] = [{"campaign": campaign, "formato": fmt, "amount": inv["subtotal"], "advertiser": inv["anunciante"]}]
    return inv


# ── File copy / rename ─────────────────────────────────────────────────────────

def classify_copy(invoices: list) -> None:
    REPORTES.mkdir(exist_ok=True)
    (ENTRADA / "sin_clasificar").mkdir(exist_ok=True)

    for inv in invoices:
        plat = inv["plataforma"]
        sub  = inv.get("subcategoria", "")
        num  = inv["numero"]
        adv  = inv["anunciante"]
        src: Path = inv["source_path"]

        if plat == "Google" and sub == "DV360":
            dest_dir = PROCESADAS / "Google" / "DV360"
            new_name = f"{num}_Google_DV360_{adv}.pdf"
        elif plat == "Google" and sub == "CM360":
            dest_dir = PROCESADAS / "Google" / "CM360"
            new_name = f"{num}_Google_CM360_{adv}.pdf"
        elif plat == "Meta":
            dest_dir = PROCESADAS / "Meta"
            new_name = f"{num}_Meta_{adv}.pdf"
        elif plat == "TikTok":
            dest_dir = PROCESADAS / "TikTok"
            new_name = f"{num}_TikTok_{adv}.pdf"
        else:
            dest_dir = ENTRADA / "sin_clasificar"
            new_name = src.name

        dest_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(src, dest_dir / new_name)
        inv["dest_name"] = new_name


# ── Excel styles ───────────────────────────────────────────────────────────────

HDR_FILL    = PatternFill("solid", fgColor="1F4E79")
HDR_FONT    = Font(bold=True, color="FFFFFF")
GOOGLE_FILL = PatternFill("solid", fgColor="E2EFDA")  # verde claro
META_FILL   = PatternFill("solid", fgColor="D6E4F0")  # azul claro
TIKTOK_FILL = PatternFill("solid", fgColor="FFD9E8")  # rosa claro
SUB_FILL    = PatternFill("solid", fgColor="FFF2CC")
TOT_FILL    = PatternFill("solid", fgColor="FCE4D6")
BOLD        = Font(bold=True)
NUM_FMT     = "#,##0.00"
PLAT_FILL   = {"Google": GOOGLE_FILL, "Meta": META_FILL, "TikTok": TIKTOK_FILL}


def set_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ── Excel generation ───────────────────────────────────────────────────────────

def write_excel(invoices: list, out: Path) -> None:
    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen por Plataforma ────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen por Plataforma"
    ws1.append(["N° Factura", "Plataforma", "Subcategoría", "Anunciante",
                "Costo de Medios (ARS)", "Tech Fee (ARS)", "Subtotal (ARS)", "IIBB Meta (ARS)"])
    set_header(ws1)
    ws1.freeze_panes = "A2"
    ws1.row_dimensions[1].height = 28

    grand_total = 0.0
    for plat in ["Google", "Meta", "TikTok"]:
        plat_invs = [i for i in invoices if i["plataforma"] == plat]
        if not plat_invs:
            continue
        plat_total = 0.0
        fill = PLAT_FILL[plat]

        for inv in plat_invs:
            # Use pre-computed totals if available (set after Spent enrichment)
            costo_medios = inv.get("costo_medios_total") or sum(
                l["amount"] for l in inv.get("lineas", []) if l["formato"] == "Costo de medios"
            )
            tech_fee = inv.get("tech_fee_total") or sum(
                l["amount"] for l in inv.get("lineas", []) if l["formato"] == "Tech Fee"
            )
            sub_display = {"CM360": "Adserver"}.get(inv.get("subcategoria", ""), inv.get("subcategoria", ""))
            iibb = inv.get("iibb") or None
            ws1.append([
                inv["numero"], plat, sub_display, inv["anunciante"],
                costo_medios or None, tech_fee or None, inv["subtotal"], iibb,
            ])
            r = ws1.max_row
            for c in range(1, 9):
                ws1.cell(r, c).fill = fill
            for c in (5, 6, 7, 8):
                ws1.cell(r, c).number_format = NUM_FMT
            plat_total  += inv["subtotal"]
            grand_total += inv["subtotal"]

            # Sub-filas de formato para DV360 (del Spent file) con Tech Fee por formato
            if inv.get("subcategoria") == "DV360" and inv.get("lineas"):
                for line in inv["lineas"]:
                    tf = format_tech_fee(line["formato"], line["amount"])
                    net = round(line["amount"] - tf, 2)  # inversión real en plataforma
                    ws1.append([
                        "", "", f"  ↳ {line['formato']}", "",
                        None, tf, net, None,
                    ])
                    r = ws1.max_row
                    for c in range(1, 9):
                        ws1.cell(r, c).fill = fill
                    ws1.cell(r, 6).number_format = NUM_FMT
                    ws1.cell(r, 7).number_format = NUM_FMT
                    ws1.cell(r, 3).font = Font(italic=True, color="3A6E3A")
                    ws1.cell(r, 6).font = Font(italic=True)
                    ws1.cell(r, 7).font = Font(italic=True)

        ws1.append([f"SUBTOTAL {plat.upper()}", "", "", "", None, None, plat_total, None])
        r = ws1.max_row
        for c in range(1, 9):
            ws1.cell(r, c).fill = SUB_FILL
            ws1.cell(r, c).font = BOLD
        ws1.cell(r, 7).number_format = NUM_FMT

    ws1.append(["TOTAL GENERAL", "", "", "", None, None, grand_total, None])
    r = ws1.max_row
    for c in range(1, 9):
        ws1.cell(r, c).fill = TOT_FILL
        ws1.cell(r, c).font = BOLD
    ws1.cell(r, 7).number_format = NUM_FMT

    ws1.column_dimensions["A"].width = 24
    ws1.column_dimensions["B"].width = 12
    ws1.column_dimensions["C"].width = 16
    ws1.column_dimensions["D"].width = 22
    ws1.column_dimensions["E"].width = 22
    ws1.column_dimensions["F"].width = 18
    ws1.column_dimensions["G"].width = 22
    ws1.column_dimensions["H"].width = 20

    # ── Hoja 2: Detalle por Campaña ───────────────────────────────────
    ws2 = wb.create_sheet("Detalle por Campaña")
    ws2.append(["N° Factura", "Plataforma", "Anunciante", "Campaña", "Formato", "Subtotal Línea (ARS)"])
    set_header(ws2)
    ws2.freeze_panes = "A2"
    ws2.row_dimensions[1].height = 28

    for inv in invoices:
        plat = inv["plataforma"]
        fill = PLAT_FILL.get(plat)
        for line in inv.get("lineas", []):
            ws2.append([
                inv["numero"], plat, line["advertiser"],
                line["campaign"], line["formato"], line["amount"],
            ])
            r = ws2.max_row
            for c in range(1, 7):
                if fill:
                    ws2.cell(r, c).fill = fill
            ws2.cell(r, 6).number_format = NUM_FMT

    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 65
    ws2.column_dimensions["E"].width = 28
    ws2.column_dimensions["F"].width = 22

    # ── Hoja 3: Sin Clasificar (si aplica) ────────────────────────────
    unclass = [i for i in invoices if i["plataforma"] == "Desconocido"]
    if unclass:
        ws3 = wb.create_sheet("Sin Clasificar")
        ws3.append(["Archivo"])
        set_header(ws3)
        for inv in unclass:
            ws3.append([inv["source_path"].name])

    wb.save(out)


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    invoices    = []
    unclassified = []

    def add_unclassified(pdf):
        unclassified.append({
            "plataforma": "Desconocido", "source_path": pdf,
            "numero": pdf.stem, "anunciante": "?",
            "subtotal": 0.0, "subcategoria": "", "lineas": [],
        })

    # ── Root PDFs (Meta + TikTok) ─────────────────────────────────────
    for pdf in sorted(ENTRADA.glob("*.pdf")):
        try:
            with pdfplumber.open(pdf) as p:
                first = (p.pages[0].extract_text() or "") if p.pages else ""

            if "Meta Platforms" in first:
                inv = parse_meta(pdf)
            elif "TikTok" in first or "IMS Argentina" in first or "Aleph" in first:
                inv = parse_tiktok(pdf)
            else:
                print(f"  ⚠ No clasificado: {pdf.name}")
                add_unclassified(pdf)
                continue

            invoices.append(inv)
            print(f"  ✓ {pdf.name}")
            print(f"    → {inv['plataforma']} | {inv['anunciante']} | ARS {inv['subtotal']:>18,.2f}")
        except Exception as e:
            print(f"  ✗ Error {pdf.name}: {e}")
            add_unclassified(pdf)

    # ── Google PDFs ───────────────────────────────────────────────────
    google_dir = ENTRADA / "Google"
    if google_dir.exists():
        for pdf in sorted(google_dir.glob("*.pdf")):
            try:
                inv = parse_google(pdf)
                invoices.append(inv)
                print(f"  ✓ {pdf.name}")
                print(f"    → Google {inv['subcategoria']} | {inv['anunciante']} | ARS {inv['subtotal']:>15,.2f}")
            except Exception as e:
                print(f"  ✗ Error {pdf.name}: {e}")
                add_unclassified(pdf)

    # ── Enriquecer DV360 con detalle de formato (Spent file) ─────────
    spent_data = load_spent_data()
    for inv in invoices:
        if inv.get("subcategoria") == "DV360":
            # Preserve Costo de medios and Tech Fee totals for the Resumen columns
            inv["costo_medios_total"] = sum(
                l["amount"] for l in inv.get("lineas", []) if l["formato"] == "Costo de medios"
            )
            inv["tech_fee_total"] = sum(
                l["amount"] for l in inv.get("lineas", []) if l["formato"] == "Tech Fee"
            )
            # Replace lineas with per-format Spent breakdown if available
            adv = inv["anunciante"]
            if adv in spent_data:
                inv["lineas"] = spent_data[adv]
                print(f"    → Detalle Spent: {[l['formato'] for l in inv['lineas']]}")

    all_invoices = invoices + unclassified

    # ── Copy & rename ─────────────────────────────────────────────────
    classify_copy(all_invoices)

    # ── Excel ─────────────────────────────────────────────────────────
    excel_path = REPORTES / "Resumen_Facturas_Abril_2026.xlsx"
    write_excel(all_invoices, excel_path)

    # ── Summary ───────────────────────────────────────────────────────
    dv360   = sum(1 for i in invoices if i.get("subcategoria") == "DV360")
    cm360   = sum(1 for i in invoices if i.get("subcategoria") == "CM360")
    meta    = sum(1 for i in invoices if i["plataforma"] == "Meta")
    tiktok  = sum(1 for i in invoices if i["plataforma"] == "TikTok")
    unclass_n = len(unclassified)

    print(f"\n✅ Procesados: {len(invoices)} archivos")
    print(f"📁 Google DV360: {dv360} | Google CM360: {cm360} | Meta: {meta} | TikTok: {tiktok}")
    print(f"⚠️  Sin clasificar: {unclass_n}")
    print(f"💾 Excel guardado en: {excel_path}")


if __name__ == "__main__":
    main()
